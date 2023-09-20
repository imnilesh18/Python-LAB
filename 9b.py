import openpyxl

print('Opening workbook...')
wb = openpyxl.load_workbook('IA_MARKS.xlsx', data_only=True)
# data_only is required to read only the cell value and not the formula
sheet = wb['book1']
print(" Rows in the workbook:", sheet.max_row)
print('Reading rows...')
for row in range(1, sheet.max_row + 1):
    USN = sheet['A' + str(row)].value
    Name = sheet['B' + str(row)].value
    IAT1 = sheet['C' + str(row)].value
    IAT2 = sheet['D' + str(row)].value
    IAT3 = sheet['E' + str(row)].value
    Total = sheet['F' + str(row)].value
    Avg = sheet['G' + str(row)].value
    print("%4s %12s %20s %7s %7s %7s %7s %7.2s" \
          % (row, USN, Name, IAT1, IAT2, IAT3, Total, Avg))

# now writing/appending values - Use loop to add more
row += 1
USN = input("Enter USN : ")
Name = input("Enter Name : ")
IAT1 = int(input("Enter marks in IAT1 : "))
IAT2 = int(input("Enter marks in IAT2 : "))
IAT3 = int(input("Enter marks in IAT3 : "))
Total = IAT1 + IAT2 + IAT3
Avg = Total / 3
sheet['A' + str(row)] = USN
sheet['B' + str(row)] = Name
sheet['C' + str(row)] = IAT1
sheet['D' + str(row)] = IAT2
sheet['E' + str(row)] = IAT2
sheet['F' + str(row)] = Total
sheet['G' + str(row)] = Avg
wb.save('IA_MARKS.xlsx')
wb.close()
# reopening to check whether record is appended/written to the workbook
wb = openpyxl.load_workbook('IA_MARKS.xlsx', data_only=True)
# data_only is required to read only the cell value and not the formula
sheet = wb['book1']
# we can open active sheet by default : sheet = wb.active
print('Reading rows...')
for row in range(1, sheet.max_row + 1):
    USN = sheet['A' + str(row)].value
    Name = sheet['B' + str(row)].value
    IAT1 = sheet['C' + str(row)].value
    IAT2 = sheet['D' + str(row)].value
    IAT3 = sheet['E' + str(row)].value
    Total = sheet['F' + str(row)].value
    Avg = sheet['G' + str(row)].value
    print("%4s %12s %20s %7s %7s %7s %7s %7.2s" \
          % (row, USN, Name, IAT1, IAT2, IAT3, Total, Avg))
wb.close()
